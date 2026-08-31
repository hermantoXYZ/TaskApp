import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.urls import reverse
from django.views.generic import TemplateView
from web_project import TemplateLayout
from django.contrib import messages
from .models import (
    Course, CourseParticipant, CourseAgenda, CourseAssignment, 
    CourseAttendance, StudentAssignmentSubmission, CourseQuiz, 
    StudentQuizAttempt
)
from .decorators_dosen import DosenRequiredMixin

class AcademyView(TemplateView):
    def get_context_data(self, **kwargs):
        context = TemplateLayout.init(self, super().get_context_data(**kwargs))
        return context

class CourseRecapitulationView(DosenRequiredMixin, AcademyView):
    template_name = "course_rekapitulasi.html" 

    def get(self, request, course_uuid, *args, **kwargs):
        course = get_object_or_404(Course, uuid=course_uuid)
        participants = CourseParticipant.objects.filter(course=course)\
            .select_related('mahasiswa', 'mahasiswa__prodi')\
            .prefetch_related('group_memberships__group')\
            .order_by('mahasiswa__nim')

        agendas = CourseAgenda.objects.filter(course=course, is_active=True).order_by('session_number', 'agenda_date')
        all_assignments = CourseAssignment.objects.filter(agenda__course=course, agenda__is_active=True).order_by('due_date')
        assignments_individu = [t for t in all_assignments if t.assignment_type != 'group']
        assignments_group = [t for t in all_assignments if t.assignment_type == 'group']

        quizzes = CourseQuiz.objects.filter(course=course).order_by('created_at') 
        total_agendas = agendas.count()
        POINTS_MAP = {'present': 100, 'late': 75, 'sick': 60, 'excused': 50, 'absent': 0, '-': 0}
        all_attendances = CourseAttendance.objects.filter(agenda__course=course, agenda__is_active=True)
        attendance_map = {(att.participant_id, att.agenda_id): att.status for att in all_attendances}
        all_submissions = StudentAssignmentSubmission.objects.filter(assignment__agenda__course=course, assignment__agenda__is_active=True)
        submission_map = {(sub.assignment.id, sub.student.nim): sub for sub in all_submissions}
        rekap_data = [] 

        for p in participants:
            student_agenda_statuses = []
            current_total_points = 0 
            
            for ag in agendas:
                status = attendance_map.get((p.id, ag.id), '-') 
                student_agenda_statuses.append({'agenda_id': ag.id, 'status': status})
                current_total_points += POINTS_MAP.get(status, 0)
            
            attendance_score = 0
            if total_agendas > 0:
                max_possible_points = total_agendas * 100
                attendance_score = round((current_total_points / max_possible_points) * 100, 1)

            def get_grades_list(task_list, student_obj):
                grades_result = []
                for task in task_list:
                    sub = submission_map.get((task.id, student_obj.nim))
                    
                    score = 0
                    status = 'missing'
                    sub_link = "-"
                    sub_time = "-"
                    
                    if sub:
                        sub_link = sub.submitted_link if sub.submitted_link else "-"
                        if sub.submitted_at:
                            local_time = timezone.localtime(sub.submitted_at)
                            sub_time = local_time.strftime("%d/%m %H:%M")

                        if sub.score is not None:
                            score = sub.score
                            status = 'graded'
                        else:
                            status = 'submitted'
                    
                    grades_result.append({
                        'task_id': task.id,
                        'assignment_type': task.assignment_type,
                        'score': score,
                        'status': status,
                        'link': sub_link,
                        'time': sub_time
                    })
                
                return grades_result

            grades_individu = get_grades_list(assignments_individu, p.mahasiswa)
            grades_group = get_grades_list(assignments_group, p.mahasiswa)

            student_quiz_grades = []
            for quiz in quizzes:
                attempt = StudentQuizAttempt.objects.filter(
                    quiz=quiz, participant=p, finished_at__isnull=False
                ).order_by('-total_score').first()

                student_quiz_grades.append({
                    'id': quiz.id,
                    'score': attempt.total_score if attempt else None,
                    'is_finished': True if attempt else False
                })

            rekap_data.append({
                'participant': p,
                'agenda_statuses': student_agenda_statuses,
                'attendance_score': attendance_score, 
                'grades_individu': grades_individu, 
                'grades_group': grades_group,
                'quiz_grades': student_quiz_grades,
            })

        # Cek Export Excel
        if request.GET.get('export') == 'excel':
            return self.export_to_excel(request, course, agendas, assignments_individu, assignments_group, quizzes, rekap_data)

        return self.render_to_response(self.get_context_data(
            course=course,
            agendas=agendas,
            assignments_individu=assignments_individu,
            assignments_group=assignments_group,
            quizzes=quizzes,
            rekap_data=rekap_data, 
            total_agendas=total_agendas,
        ))

    def export_to_excel(self, request, course, agendas, assignments_ind, assignments_grp, quizzes, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rekapitulasi Kelas"

        # --- STYLING ---
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        ind_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid") 
        grp_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        quiz_fill = PatternFill(start_color="FFE0DB", end_color="FFE0DB", fill_type="solid")
        link_font = Font(color="0000FF", underline="single", bold=True)

        # --- HEADER ROW 1 ---
        headers = ["No", "Nama Mahasiswa", "NIM", "Prodi", "Kelas", "Kelompok", "Kode MK"] 
        agenda_start_col = len(headers) + 1
        
        # Absensi
        for i, ag in enumerate(agendas):
            headers.append(f"P-{ag.session_number if ag.session_number is not None else i+1}")
        headers.append("Skor Absen")

        # Tugas Individu
        for i, task in enumerate(assignments_ind):
            prefix = f"Ind-{i+1}"
            headers.extend([f"{prefix} Link", f"{prefix} Waktu", f"{prefix} Nilai"])
        
        # Tugas Kelompok
        for i, task in enumerate(assignments_grp):
            prefix = f"Grp-{i+1}"
            headers.extend([f"{prefix} Link", f"{prefix} Waktu", f"{prefix} Nilai"])

        # Kuis / Ujian
        for i, quiz in enumerate(quizzes):
            q_type = getattr(quiz, 'quiz_type', 'Quiz')
            headers.append(str(q_type).upper())

        ws.append(headers)

        # --- HEADER COLORS ---
        start_ind = 7 + len(agendas) + 1 + 1
        end_ind = start_ind + (len(assignments_ind) * 3) - 1
        
        start_grp = end_ind + 1
        end_grp = start_grp + (len(assignments_grp) * 3) - 1
        
        start_quiz = end_grp + 1
        end_quiz = start_quiz + len(quizzes) - 1

        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border
            if start_ind <= cell.col_idx <= end_ind:
                cell.fill = ind_fill
            elif start_grp <= cell.col_idx <= end_grp:
                cell.fill = grp_fill
            elif start_quiz <= cell.col_idx <= end_quiz:
                cell.fill = quiz_fill
            else:
                cell.fill = header_fill

        scheme = request.scheme
        host = request.get_host()
        base_url = f"{scheme}://{host}"

        for i, ag in enumerate(agendas):
            col_idx = agenda_start_col + i
            cell = ws.cell(row=1, column=col_idx)
            first_material = ag.materials.filter(is_published=True).order_by('order').first()
            first_assignment = ag.assignments.filter(is_published=True).order_by('due_date').first()

            relative_path = ""
            p_label = f"P-{ag.session_number if ag.session_number is not None else i+1}"

            try:
                if first_material:
                    relative_path = reverse('course-preview-public-material', kwargs={
                        'course_uuid': course.uuid, 
                        'material_id': first_material.id
                    })
                    
                elif first_assignment:
                    relative_path = reverse('course-preview-public-assignment', kwargs={
                        'course_uuid': course.uuid, 
                        'assignment_id': first_assignment.id
                    })
                    
                else:
                    relative_path = reverse('course-preview-public', kwargs={
                        'course_uuid': course.uuid
                    })

                full_link = f"{base_url}{relative_path}"
                
                # Set Hyperlink di Excel
                cell.hyperlink = full_link
                cell.font = link_font  
                cell.value = p_label 
                
            except Exception as e:
                cell.value = p_label


        # --- ISI DATA ---
        for idx, item in enumerate(data):
            mhs = item['participant'].mahasiswa
            nim_str = mhs.nim.username if mhs else "NIM Tidak Ditemukan"
            nama_str = f"{mhs.nim.first_name}".strip() if mhs else "Tanpa Nama"
            prodi_str = mhs.prodi.nama_prodi if mhs and mhs.prodi else "-"
            
            group_name = "-"
            group_member = item['participant'].group_memberships.first()
            if group_member:
                group_name = group_member.group.name

            row = [idx + 1, nama_str, nim_str, prodi_str, course.group, group_name, course.code]

            # Absensi
            for stat in item['agenda_statuses']:
                code_map = {'present': 'H', 'late': 'T', 'absent': 'A', 'sick': 'S', 'excused': 'I', '-': '-'}
                row.append(code_map.get(stat['status'], '-'))
            row.append(item['attendance_score'])

            # Tugas Individu
            for grade in item['grades_individu']:
                row.extend([grade['link'], grade['time'], grade['score'] or 0])

            # Tugas Kelompok
            for grade in item['grades_group']:
                row.extend([grade['link'], grade['time'], grade['score'] or 0])
            
            # Kuis / Ujian
            for q_grade in item['quiz_grades']:
                row.append(q_grade['score'] or 0)

            ws.append(row)
            
            # --- HYPERLINK LINK TUGAS ---
            current_row = ws.max_row
            
            def apply_links(grade_list, start_col):
                for k, grade in enumerate(grade_list):
                    url = grade['link']
                    if url and url != "-" and str(url).startswith('http'):
                        c_idx = start_col + (k * 3)
                        cell = ws.cell(row=current_row, column=c_idx)
                        cell.value = "Link"       
                        cell.hyperlink = url       
                        cell.font = link_font     
                        cell.alignment = center_align
            
            apply_links(item['grades_individu'], start_ind)
            apply_links(item['grades_group'], start_grp)

        for col in ws.columns:
            if "Link" in str(col[0].value):
                ws.column_dimensions[col[0].column_letter].width = 12
            else:
                ws.column_dimensions[col[0].column_letter].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Rekap_Lengkap_{course.code}.xlsx'
        wb.save(response)
        return response