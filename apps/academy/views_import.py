import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.contrib import messages
from django.db import transaction
from django.views import View

from web_project import TemplateLayout
from django.views.generic import TemplateView

from .decorators_prodi import ProdiRequiredMixin
from .models import Course, CoursePeriod, CourseParticipant, UserDosen, UserMhs, Prodi

class ImportBaseView(TemplateView):
    def get_context_data(self, **kwargs):
        return TemplateLayout.init(self, super().get_context_data(**kwargs))

HEADER_FILL   = PatternFill("solid", fgColor="1976D2")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
EXAMPLE_FILL  = PatternFill("solid", fgColor="E3F2FD")
EXAMPLE_FONT  = Font(color="1565C0", italic=True, size=10)
THIN_BORDER   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def _make_wb(sheet_name, headers, examples):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill       = HEADER_FILL
        cell.font       = HEADER_FONT
        cell.alignment  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border     = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Baris contoh
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = EXAMPLE_FILL
            cell.font   = EXAMPLE_FONT
            cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30
    return wb


def _wb_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

class SetupSemesterView(ProdiRequiredMixin, ImportBaseView):
    template_name = "prodi/setup/setup_semester.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_periods']  = CoursePeriod.objects.filter(is_active=True).order_by('-start_date')
        context['active_period'] = CoursePeriod.objects.filter(is_active=True).first()
        context['all_prodis']   = Prodi.objects.all().order_by('nama_prodi')
        return context 


# ─────────────────────────────────────────────────────────────────────────────
# 2. Download Template Excel
# ─────────────────────────────────────────────────────────────────────────────
TEMPLATES_CONFIG = {
    'courses': {
        'filename': 'template_matakuliah.xlsx',
        'sheet':    'Mata Kuliah',
        'headers':  ['kode_mk', 'nama_mk', 'deskripsi', 'sks_t', 'sks_p', 'kelas', 'link_rps'],
        'examples': [
            ['TIK101', 'Algoritma & Pemrograman', 'Dasar pemrograman', 3, 1, 'A', ''],
            ['TIK102', 'Basis Data', 'Perancangan database', 2, 2, 'B', 'https://rps.kampus.ac.id/tik102'],
        ],
    },
    'coaches': {
        'filename': 'template_dosen_mk.xlsx',
        'sheet':    'Dosen MK',
        'headers':  ['kode_mk', 'kelas', 'nip_dosen'],
        'examples': [
            # nip_dosen: 1 NIP atau beberapa NIP dipisah koma
            ['TIK101', 'A', '198001012005011001,197512152003121002'],  # 2 dosen sekaligus
            ['TIK101', 'B', '198001012005011001'],                     # 1 dosen
            ['TIK102', 'A', '198001012005011001,197512152003121002,198811051999031001'],  # 3 dosen
        ],
    },
    'participants': {
        'filename': 'template_peserta_mk.xlsx',
        'sheet':    'Peserta MK',
        'headers':  ['kode_mk', 'kelas', 'nim_mahasiswa'],
        'examples': [
            ['TIK101', 'A', '22310001'],
            ['TIK101', 'A', '22310002'],
            ['TIK102', 'B', '22310001'],
        ],
    },
}


class DownloadTemplateView(ProdiRequiredMixin, View):
    def get(self, request, tipe, *args, **kwargs):
        cfg = TEMPLATES_CONFIG.get(tipe)
        if not cfg:
            messages.error(request, 'Tipe template tidak valid.')
            return redirect('setup-semester')

        wb = _make_wb(cfg['sheet'], cfg['headers'], cfg['examples'])
        return _wb_response(wb, cfg['filename'])

def _read_excel(uploaded_file, expected_headers):
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        ws = wb.active
    except Exception as e:
        return None, f"File tidak dapat dibaca: {e}"

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return None, "File kosong."

    # Normalisasi header
    actual = [str(h).strip().lower() if h else '' for h in header_row]
    missing = [h for h in expected_headers if h not in actual]
    if missing:
        return None, f"Kolom wajib tidak ditemukan: {', '.join(missing)}"

    col_idx = {h: actual.index(h) for h in expected_headers}

    rows = []
    for row in rows_iter:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue  # skip baris kosong
        entry = {h: (row[col_idx[h]] if row[col_idx[h]] is not None else '') for h in expected_headers}
        rows.append(entry)

    return rows, None

def _to_str(val):
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()

class ImportCoursesView(ProdiRequiredMixin, View):
    HEADERS = ['kode_mk', 'nama_mk', 'deskripsi', 'sks_t', 'sks_p', 'kelas', 'link_rps']

    def post(self, request, *args, **kwargs):
        file     = request.FILES.get('file_courses')
        period_id = request.POST.get('period_id', '').strip()
        prodi_id  = request.POST.get('prodi_id', '').strip()

        errors = []

        # Validasi file
        if not file:
            messages.error(request, 'File Excel mata kuliah tidak ditemukan.')
            return redirect('setup-semester')

        # Validasi periode
        try:
            period = CoursePeriod.objects.get(id=period_id)
        except (CoursePeriod.DoesNotExist, Exception):
            messages.error(request, 'Periode tidak valid. Silakan pilih periode terlebih dahulu.')
            return redirect('setup-semester')

        # Validasi prodi
        try:
            from .models import UserProdi
            user_prodi = UserProdi.objects.get(username=request.user)
            prodi = user_prodi.prodi
        except Exception:
            prodi = None

        if prodi_id and not prodi:
            try:
                prodi = Prodi.objects.get(id=prodi_id)
            except Prodi.DoesNotExist:
                messages.error(request, 'Program studi tidak valid.')
                return redirect('setup-semester')

        if not prodi:
            messages.error(request, 'Tidak dapat menentukan Program Studi. Pastikan akun Anda terikat ke Prodi.')
            return redirect('setup-semester')

        # Baca Excel
        rows, read_error = _read_excel(file, self.HEADERS)
        if read_error:
            messages.error(request, f'Error membaca file: {read_error}')
            return redirect('setup-semester')

        if not rows:
            messages.warning(request, 'File Excel tidak memiliki data.')
            return redirect('setup-semester')

        success_count = 0
        skip_count    = 0

        try:
            with transaction.atomic():
                for i, row in enumerate(rows, start=2):
                    kode  = str(row['kode_mk']).strip()
                    nama  = str(row['nama_mk']).strip()
                    kelas = str(row['kelas']).strip()

                    if not kode or not nama or not kelas:
                        errors.append(f"Baris {i}: kode_mk, nama_mk, dan kelas wajib diisi.")
                        continue

                    # Validasi SKS
                    try:
                        sks_t = int(row['sks_t'])
                        sks_p = int(row['sks_p'])
                    except (ValueError, TypeError):
                        errors.append(f"Baris {i} [{kode}]: sks_t dan sks_p harus berupa angka.")
                        continue

                    # Cek duplikat: kode + kelas + periode
                    if Course.objects.filter(code=kode, group=kelas, period=period, prodi=prodi).exists():
                        skip_count += 1
                        continue

                    Course.objects.create(
                        code        = kode,
                        name        = nama,
                        description = str(row['deskripsi']).strip()[:500],
                        credit_t    = sks_t,
                        credit_p    = sks_p,
                        group       = kelas,
                        period      = period,
                        prodi       = prodi,
                        link_rps    = str(row['link_rps']).strip() or None,
                        is_active   = True,
                    )
                    success_count += 1

        except Exception as e:
            messages.error(request, f'Terjadi kesalahan saat menyimpan data: {e}')
            return redirect('setup-semester')

        # Ringkasan
        if success_count:
            messages.success(request, f'✅ {success_count} Mata Kuliah berhasil diimport ke periode "{period.name}".')
        if skip_count:
            messages.info(request, f'ℹ️ {skip_count} baris dilewati (sudah ada di database).')
        if errors:
            for err in errors[:10]:   # batasi 10 error agar tidak overflow
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f'… dan {len(errors) - 10} error lainnya.')

        return redirect('setup-semester')

class ImportCoachesView(ProdiRequiredMixin, View):
    """
    POST params:
      - file_coaches   : file .xlsx
      - period_id      : UUID CoursePeriod (untuk filter MK)
    """

    HEADERS = ['kode_mk', 'kelas', 'nip_dosen']

    def post(self, request, *args, **kwargs):
        file      = request.FILES.get('file_coaches')
        period_id = request.POST.get('period_id', '').strip()

        if not file:
            messages.error(request, 'File Excel dosen tidak ditemukan.')
            return redirect('setup-semester')

        try:
            period = CoursePeriod.objects.get(id=period_id)
        except (CoursePeriod.DoesNotExist, Exception):
            messages.error(request, 'Periode tidak valid.')
            return redirect('setup-semester')

        rows, read_error = _read_excel(file, self.HEADERS)
        if read_error:
            messages.error(request, f'Error membaca file: {read_error}')
            return redirect('setup-semester')

        if not rows:
            messages.warning(request, 'File Excel tidak memiliki data.')
            return redirect('setup-semester')

        success_count = 0
        skip_count    = 0
        errors        = []

        try:
            with transaction.atomic():
                for i, row in enumerate(rows, start=2):
                    kode  = str(row['kode_mk']).strip()
                    kelas = str(row['kelas']).strip()
                    nip_raw = str(row['nip_dosen']).strip()

                    if not kode or not kelas or not nip_raw:
                        errors.append(f"Baris {i}: semua kolom wajib diisi.")
                        continue

                    # Cari MK
                    course_qs = Course.objects.filter(code=kode, group=kelas, period=period)
                    if not course_qs.exists():
                        errors.append(f"Baris {i}: MK '{kode}' kelas '{kelas}' tidak ditemukan di periode '{period.name}'.")
                        continue

                    # Support multi-NIP: pisahkan dengan koma
                    # Contoh: "NIP1,NIP2,NIP3" atau cukup "NIP1"
                    nip_list = [n.strip() for n in nip_raw.split(',') if n.strip()]

                    for nip in nip_list:
                        # Cari dosen
                        try:
                            dosen = UserDosen.objects.get(nip__username=nip)
                        except UserDosen.DoesNotExist:
                            errors.append(f"Baris {i}: NIP/username dosen '{nip}' tidak ditemukan.")
                            continue

                        for course in course_qs:
                            if dosen in course.coaches.all():
                                skip_count += 1
                            else:
                                course.coaches.add(dosen)
                                success_count += 1

        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {e}')
            return redirect('setup-semester')

        if success_count:
            messages.success(request, f'✅ {success_count} penugasan dosen berhasil disimpan.')
        if skip_count:
            messages.info(request, f'ℹ️ {skip_count} penugasan dilewati (sudah terdaftar).')
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f'… dan {len(errors) - 10} error lainnya.')

        return redirect('setup-semester')

class ImportParticipantsView(ProdiRequiredMixin, View):
    """
    POST params:
      - file_participants : file .xlsx
      - period_id         : UUID CoursePeriod
    """

    HEADERS = ['kode_mk', 'kelas', 'nim_mahasiswa']

    def post(self, request, *args, **kwargs):
        file      = request.FILES.get('file_participants')
        period_id = request.POST.get('period_id', '').strip()

        if not file:
            messages.error(request, 'File Excel peserta tidak ditemukan.')
            return redirect('setup-semester')

        try:
            period = CoursePeriod.objects.get(id=period_id)
        except (CoursePeriod.DoesNotExist, Exception):
            messages.error(request, 'Periode tidak valid.')
            return redirect('setup-semester')

        rows, read_error = _read_excel(file, self.HEADERS)
        if read_error:
            messages.error(request, f'Error membaca file: {read_error}')
            return redirect('setup-semester')

        if not rows:
            messages.warning(request, 'File Excel tidak memiliki data.')
            return redirect('setup-semester')

        success_count = 0
        skip_count    = 0
        errors        = []

        try:
            with transaction.atomic():
                for i, row in enumerate(rows, start=2):
                    kode = str(row['kode_mk']).strip()
                    kelas = str(row['kelas']).strip()
                    nim  = _to_str(row['nim_mahasiswa'])
 
                    if not kode or not kelas or not nim:
                        errors.append(f"Baris {i}: semua kolom wajib diisi.")
                        continue

                    # Cari MK
                    course_qs = Course.objects.filter(code=kode, group=kelas, period=period)
                    if not course_qs.exists():
                        errors.append(f"Baris {i}: MK '{kode}' kelas '{kelas}' tidak ditemukan di periode '{period.name}'.")
                        continue

                    # Cari mahasiswa
                    try:
                        mahasiswa = UserMhs.objects.get(nim__username=nim)
                    except UserMhs.DoesNotExist:
                        errors.append(f"Baris {i}: NIM '{nim}' tidak ditemukan.")
                        continue

                    for course in course_qs:
                        _, created = CourseParticipant.objects.get_or_create(
                            course=course,
                            mahasiswa=mahasiswa,
                            defaults={'is_active': True},
                        )
                        if created:
                            success_count += 1
                        else:
                            skip_count += 1

        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {e}')
            return redirect('setup-semester')

        if success_count:
            messages.success(request, f'{success_count} peserta berhasil dienroll.')
        if skip_count:
            messages.info(request, f'{skip_count} peserta dilewati (sudah terdaftar).')
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f'… dan {len(errors) - 10} error lainnya.')

        return redirect('setup-semester')
